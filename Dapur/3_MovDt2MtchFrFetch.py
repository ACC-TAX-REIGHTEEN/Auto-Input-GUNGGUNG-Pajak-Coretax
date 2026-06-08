import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_filename = "Nota_terbaru_temp.xlsx"

df = pd.read_excel(input_filename)

def perbaiki_bulan(tanggal_str):
    if not isinstance(tanggal_str, str):
        return tanggal_str
    
    kamus_bulan = {
        ' Nop ': ' Nov ', 
        ' Mei ': ' May ', 
        ' Agu ': ' Aug ', 
        ' Okt ': ' Oct ', 
        ' Des ': ' Dec '
    }
    
    for indo, inggris in kamus_bulan.items():
        if indo in tanggal_str:
            return tanggal_str.replace(indo, inggris)
    return tanggal_str

df['Tanggal'] = df['Tanggal'].apply(perbaiki_bulan)
df['Tanggal'] = pd.to_datetime(df['Tanggal'], format='%d %b %Y', errors='coerce')
df['Tanggal'] = df['Tanggal'].dt.strftime('%Y-%m-%d')

output_filename = "Nota_terbaru_up_temp.xlsx"
df.to_excel(output_filename, index=False)

wb = load_workbook(output_filename)
ws = wb.active

for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_filename)
print("--> Selesai! File tersimpan sebagai:", output_filename)